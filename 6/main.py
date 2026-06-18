import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule

# =====================================================
# COLORS
# =====================================================

DARK = "1F2937"
BLUE = "2563EB"
GREEN = "16A34A"
RED = "DC2626"
YELLOW = "EAB308"
GRAY = "F3F4F6"
WHITE = "FFFFFF"

# =====================================================
# SAMPLE DATA
# =====================================================

mobilities = pd.DataFrame([
    {
        "Status": "ACTIVE",
        "Type": "Study",
        "Flow": "IN",
        "Person": "Alevtyna Liaskovska",
        "Country": "Ukraine",
        "University": "KNUBA",
        "Coordinator": "Tereza",
        "Start": "01.02.2024",
        "End": "30.06.2024",
        "Grant": 5570,
        "Documents": "90%",
        "Visa": "DONE"
    },
    {
        "Status": "ACTIVE",
        "Type": "Teaching",
        "Flow": "OUT",
        "Person": "Pavle Dakic",
        "Country": "Montenegro",
        "University": "UCG",
        "Coordinator": "Lucia",
        "Start": "25.04.2024",
        "End": "03.05.2024",
        "Grant": 2255,
        "Documents": "100%",
        "Visa": "DONE"
    },
    {
        "Status": "PENDING",
        "Type": "Training",
        "Flow": "IN",
        "Person": "Jas Raj Subba",
        "Country": "Bhutan",
        "University": "Sherubtse College",
        "Coordinator": "Lucia",
        "Start": "21.10.2024",
        "End": "25.10.2024",
        "Grant": 1800,
        "Documents": "60%",
        "Visa": "PENDING"
    }
])

documents = pd.DataFrame([
    {
        "Name": "Alevtyna Liaskovska",
        "Application": "DONE",
        "Visa": "DONE",
        "Insurance": "DONE",
        "GA": "DONE",
        "Recognition": "PENDING",
        "Progress": "90%"
    },
    {
        "Name": "Jas Raj Subba",
        "Application": "DONE",
        "Visa": "PENDING",
        "Insurance": "DONE",
        "GA": "DONE",
        "Recognition": "MISSING",
        "Progress": "60%"
    }
])

universities = pd.DataFrame([
    {
        "University": "Sherubtse College",
        "Country": "Bhutan",
        "Coordinator": "Ivan Spanik",
        "Email": "ivan.spanik@stuba.sk",
        "Status": "ACTIVE"
    },
    {
        "University": "University of Botswana",
        "Country": "Botswana",
        "Coordinator": "Lucia",
        "Email": "lucia@stuba.sk",
        "Status": "ACTIVE"
    }
])

budget = pd.DataFrame([
    {
        "Country": "Ukraine",
        "Budget": 104415,
        "Used": 100707,
        "Remaining": 3708,
        "Usage": "96%"
    },
    {
        "Country": "Bhutan",
        "Budget": 22720,
        "Used": 22520,
        "Remaining": 200,
        "Usage": "99%"
    }
])

cancelled = pd.DataFrame([
    {
        "Person": "Ivan Spanik",
        "Country": "Laos",
        "Type": "Training",
        "Reason": "Cancelled"
    }
])

# =====================================================
# CREATE WORKBOOK
# =====================================================

wb = Workbook()
wb.remove(wb.active)

thin = Side(style="thin", color="D1D5DB")

# =====================================================
# STYLING FUNCTIONS
# =====================================================

def style_headers(ws, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)

        cell.font = Font(
            color=WHITE,
            bold=True,
            size=11
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=BLUE
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )


def style_rows(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:

            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

            cell.alignment = Alignment(
                vertical="center"
            )

            cell.font = Font(
                name="Segoe UI",
                size=11
            )


def auto_width(ws):

    for column in ws.columns:

        max_length = 0
        letter = column[0].column_letter

        for cell in column:

            try:
                length = len(str(cell.value))

                if length > max_length:
                    max_length = length

            except:
                pass

        adjusted = min(max_length + 4, 40)

        ws.column_dimensions[letter].width = adjusted


# =====================================================
# DASHBOARD
# =====================================================

dashboard = wb.create_sheet("Dashboard")

dashboard["A1"] = "ERASMUS / ICM DASHBOARD"

dashboard["A1"].font = Font(
    size=22,
    bold=True,
    color=BLUE
)

cards = [
    ["Total Mobilities", 134],
    ["Countries", 11],
    ["Used Budget", "380 026 €"],
    ["Cancelled", 5],
    ["Pending Visas", 7],
    ["Pending Documents", 12]
]

row_num = 4

for item in cards:

    dashboard[f"A{row_num}"] = item[0]
    dashboard[f"B{row_num}"] = item[1]

    dashboard[f"A{row_num}"].font = Font(
        bold=True,
        color=WHITE
    )

    dashboard[f"A{row_num}"].fill = PatternFill(
        "solid",
        fgColor=DARK
    )

    dashboard[f"B{row_num}"].fill = PatternFill(
        "solid",
        fgColor=GRAY
    )

    dashboard[f"B{row_num}"].font = Font(
        bold=True,
        size=12
    )

    dashboard[f"A{row_num}"].alignment = Alignment(horizontal="center")
    dashboard[f"B{row_num}"].alignment = Alignment(horizontal="center")

    row_num += 2

dashboard.column_dimensions["A"].width = 30
dashboard.column_dimensions["B"].width = 20

# =====================================================
# TABLE CREATOR
# =====================================================

def create_sheet(sheet_name, dataframe):

    ws = wb.create_sheet(sheet_name)

    for row in dataframe_to_rows(
        dataframe,
        index=False,
        header=True
    ):
        ws.append(row)

    style_headers(ws, dataframe.shape[1])
    style_rows(ws)

    # alternating rows

    for row in range(2, ws.max_row + 1):

        if row % 2 == 0:

            for col in range(1, ws.max_column + 1):

                ws.cell(row=row, column=col).fill = PatternFill(
                    "solid",
                    fgColor="FAFAFA"
                )

    auto_width(ws)

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    # create excel table

    end_col = chr(64 + ws.max_column)

    table = Table(
        displayName=f"{sheet_name}Table",
        ref=f"A1:{end_col}{ws.max_row}"
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False
    )

    table.tableStyleInfo = style

    ws.add_table(table)

    return ws


# =====================================================
# CREATE SHEETS
# =====================================================

mob_ws = create_sheet("Mobilities", mobilities)
doc_ws = create_sheet("Documents", documents)
uni_ws = create_sheet("Universities", universities)
budget_ws = create_sheet("Budget", budget)
cancel_ws = create_sheet("Cancelled", cancelled)

# =====================================================
# CONDITIONAL FORMATTING
# =====================================================

green_fill = PatternFill(
    bgColor="DCFCE7"
)

yellow_fill = PatternFill(
    bgColor="FEF9C3"
)

red_fill = PatternFill(
    bgColor="FEE2E2"
)

mob_ws.conditional_formatting.add(
    "A2:A1000",
    FormulaRule(
        formula=['A2="ACTIVE"'],
        fill=green_fill
    )
)

mob_ws.conditional_formatting.add(
    "A2:A1000",
    FormulaRule(
        formula=['A2="PENDING"'],
        fill=yellow_fill
    )
)

mob_ws.conditional_formatting.add(
    "A2:A1000",
    FormulaRule(
        formula=['A2="CANCELLED"'],
        fill=red_fill
    )
)

# =====================================================
# DOCUMENT COLORS
# =====================================================

for row in doc_ws.iter_rows(min_row=2):

    for cell in row:

        if cell.value == "DONE":

            cell.fill = PatternFill(
                "solid",
                fgColor="DCFCE7"
            )

        elif cell.value == "PENDING":

            cell.fill = PatternFill(
                "solid",
                fgColor="FEF9C3"
            )

        elif cell.value == "MISSING":

            cell.fill = PatternFill(
                "solid",
                fgColor="FEE2E2"
            )

# =====================================================
# FINAL TOUCHES
# =====================================================

for sheet in wb.sheetnames:

    ws = wb[sheet]

    for row in ws.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical="center"
            )

# =====================================================
# SAVE
# =====================================================

wb.save("ICM_REDESIGNED_SYSTEM.xlsx")

print("===================================")
print("SYSTEM CREATED SUCCESSFULLY")
print("FILE: ICM_REDESIGNED_SYSTEM.xlsx")
print("===================================")